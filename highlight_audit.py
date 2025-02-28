import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Folder where reports are stored
report_folder = r"C:\Users\Owner\Dropbox\Program Team\Data\Data Quality Monitoring\ST24-25"

# Get all Excel files in the folder
files = [f for f in os.listdir(report_folder) if f.endswith(".xlsx")]

# Find the latest file by modification time
latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(report_folder, f)))

# Full path to the latest file
file_path = os.path.join(report_folder, latest_file)
print(f"Using latest file: {file_path}")

# Load the latest Excel file
wb = load_workbook(file_path)
ws = wb["Case Management Requirements"]  # Select the correct sheet

# Define the highlight color (yellow)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Columns to check for "No" responses (R, S, T, U)
no_columns = ["R", "S", "T", "U"]
for col in no_columns:
    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws[f"{col}{row}"]
        if cell.value == "No":
            cell.fill = highlight_fill

# Columns to check for "0" values (V, Y)
zero_columns = ["V", "Y"]
for col in zero_columns:
    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws[f"{col}{row}"]
        if cell.value == 0:
            cell.fill = highlight_fill

# Save the modified file
highlighted_file_path = os.path.join(report_folder, f"Highlighted_{latest_file}")
wb.save(highlighted_file_path)

print(f"Highlighted file saved as: {highlighted_file_path}")
